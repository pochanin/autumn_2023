import openpyxl

workbook = openpyxl.open("C:\Users\pm07\PycharmProjects\pythonProject\test.xlsx")
worksheet = workbook.active

f = open('test.xlsx', encoding='utf-8')

M = []
for line in f:
    M.append(line.rstrip().split(', '))

salary_sum = 0
for i in M:
    salary_sum += int(i[4])

M = sorted(M, key=lambda x: x[3])
for i in range(len(M)):
    worksheet.cell(row=i+1, column=1, value=M[i][0])
    worksheet.cell(row=i+1, column=2, value=M[i][1])
    worksheet.cell(row=i+1, column=3, value=M[i][2])
    worksheet.cell(row=i+1, column=4, value=M[i][3])
    worksheet.cell(row=i+1, column=5, value=M[i][4])
worksheet.cell(row=len(M)+1, column=1, value='ИТОГО:')
worksheet.cell(row=len(M)+1, column=2, value=salary_sum)

workbook.save("C:\Users\pm07\PycharmProjects\pythonProject\test.xlsx")
