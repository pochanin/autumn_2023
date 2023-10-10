import openpyxl

workbook = openpyxl.open("C:\Users\pm07\PycharmProjects\pythonProject\test.xlsx")
worksheet = workbook.active

M = []
for i in range(1, worksheet.max_row + 1):
    M.append(worksheet[i][1].value)
M = sorted(M)

worksheet_2 = workbook.worksheets[1]

worksheet_2.cell(row=1, column=1, value='Минимальное значение')
worksheet_2.cell(row=1, column=2, value=M[0])
worksheet_2.cell(row=2, column=1, value='Максимальное значение')
worksheet_2.cell(row=2, column=2, value=max(M))
worksheet_2.cell(row=3, column=1, value='Среднее арифметическое')
worksheet_2.cell(row=3, column=2, value=(sum(M)/len(M)))
worksheet_2.cell(row=4, column=1, value='Медиана')
if len(M) % 2 == 0:
    worksheet_2.cell(row=4, column=2, value=(M[len(M) // 2] + M[len(M) // 2 - 1) / 2)
else:
    worksheet_2.cell(row=4, column=2, value=(M[len(M) // 2]))
workbook.save("C:\Users\pm07\PycharmProjects\pythonProject\test.xlsx")