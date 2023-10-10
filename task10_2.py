import openpyxl

workbook = openpyxl.open("C:\Users\pm07\PycharmProjects\pythonProject\test.xlsx")
worksheet = workbook.active

dic = {}
for i in range(1, worksheet.max_row + 1):
    if worksheet[i][0].value not in dic.keys():
        dic[worksheet[i][0].value] = worksheet[i][1].value
    else:
        dic[worksheet[i][0].value] += worksheet[i][1].value


worksheet = workbook.worksheets[1]
for i in range(1, worksheet.max_row + 1):
    if worksheet[i][0].value not in dic.keys():
        dic[worksheet[i][0].value] = worksheet[i][1].value
    else:
        dic[worksheet[i][0].value] += worksheet[i][1].value

dic = sorted(dic.items())

worksheet_2 = workbook.worksheets[2]
for i in range(len(dic)):
    Item_1 = worksheet_2.cell(row=i+1, column=1)
    Item_2 = worksheet_2.cell(row=i+1, column=2)
    Item_1.value = dic[i][0]
    Item_2.value = dic[i][1]
workbook.save("C:\Users\pm07\PycharmProjects\pythonProject\test.xlsx")
